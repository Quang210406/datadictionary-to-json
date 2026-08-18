"""What happens to an Excel archive that has no cloud stage.

    python tests/test_no_cloud_stage.py

No API key, no quota, no network. Only the ONE ai-shaped step is stubbed:
agent.convert is replaced with a perfect reader of the synthetic sheets this
file writes. Everything downstream is the real thing — catalog, assembly,
validation, coverage, emit — which is exactly the code that decides what a
missing stage does. The project's rule is "AI reads files, Python joins them",
and the join is what is under test here.

Three archives, all built from the same two hops:

  A  layout declares a cloud stage, and a cloud workbook exists
     the baseline, so the other two have something to differ from
  B  layout declares a cloud stage, but no cloud workbook is present
     the realistic accident: the config says four stages, the folder has three
  C  layout declares no cloud stage at all
     the honest three-stage archive
"""

import json
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook

import assemble
import store as store_mod
from catalog import build_catalog, cloud_sheets, find_cloud_workbook
from emit import write_workbook
from layout import cloud_stage, table_prefixes, validate_layout
from store import RecordStore
from validate import chain_diagnostics, coverage_metrics

# --- the synthetic archive -------------------------------------------------
#
# Two fields carried through two hops, so a complete chain is source ->
# staging -> dwh, plus a cloud copy where one exists. Names are invented.

FIELDS = [("PARTY_ID", "VARCHAR2", "20"), ("OPEN_DT", "DATE", "8")]

HOPS = [
    # (folder, target table, source table)
    ("SRC_STG", "STG_PARTY", "RAW_PARTY"),
    ("STG_DWH", "DWH_PARTY", "STG_PARTY"),
]


def _hop_sheet(path, target, source):
    """One hop workbook: a metadata block, then a real header row, then rows.

    The metadata block matters: expected_row_count finds the header row by
    anchor words, and a stray one-cell line above it must not be mistaken for
    the header. This mirrors the shape of a real hand-maintained sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = target
    ws.append(["Target Table Name", target])
    ws.append([])
    ws.append(["Target Column", "Target Datatype", "Target Size",
               "Source Table", "Source Column", "Notes"])
    for column, datatype, size in FIELDS:
        ws.append([column, datatype, size, source, column, "1-1"])
    wb.save(path)


def _cloud_sheet(path, tables):
    """The final-stage workbook: one sheet per table, no source side."""
    wb = Workbook()
    wb.remove(wb.active)
    for table in tables:
        ws = wb.create_sheet(title=table)
        ws.append(["Cloud table", table])
        ws.append([])
        ws.append(["Column", "Datatype", "Length", "Nullable"])
        for column, datatype, size in FIELDS:
            ws.append([column, datatype, size, "Y"])
    wb.save(path)


def build_archive(root, with_cloud_workbook):
    root = Path(root)
    for folder, target, source in HOPS:
        (root / folder).mkdir(parents=True, exist_ok=True)
        _hop_sheet(root / folder / f"{target}.xlsx", target, source)
    if with_cloud_workbook:
        # The cloud sheet files the table without its warehouse prefix, the
        # same way the real one does.
        _cloud_sheet(root / "Mapping_to_CLOUD.xlsx", ["PARTY"])
    return root


LAYOUT_WITH_CLOUD = {
    "stages": ["source", "staging", "dwh", "cloud"],
    "hops": [{"dir": "SRC_STG", "from": "source", "to": "staging"},
             {"dir": "STG_DWH", "from": "staging", "to": "dwh"}],
    "cloud": {"glob": "*CLOUD*.xlsx", "stage": "cloud",
              "table_prefixes": ["DWH_"]},
    "non_hop_sheets": ["DDL"],
}

LAYOUT_NO_CLOUD = {
    "stages": ["source", "staging", "dwh"],
    "hops": [{"dir": "SRC_STG", "from": "source", "to": "staging"},
             {"dir": "STG_DWH", "from": "staging", "to": "dwh"}],
    "non_hop_sheets": ["DDL"],
}


# --- the only stub ---------------------------------------------------------

def _fake_convert(text, schema, kind):
    """A perfect agent: reads the CSV the real extractor produced.

    Returning exactly what a flawless model would return isolates the join
    logic. If a number below is wrong, assembly is wrong, not extraction.
    """
    rows = [line.split(",") for line in text.strip().splitlines()]
    out = []
    if kind == "hop_spec":
        target_table = rows[0][1].strip()
        header = next(i for i, r in enumerate(rows)
                      if r and r[0].strip().lower() == "target column")
        for row in rows[header + 1:]:
            if not row or not row[0].strip():
                continue
            out.append({
                "target_table": target_table, "target_column": row[0].strip(),
                "target_datatype": row[1].strip() or None,
                "target_size": row[2].strip() or None,
                "source_table": row[3].strip() or None,
                "source_column": row[4].strip() or None,
                "source_datatype": row[1].strip() or None,
                "source_size": row[2].strip() or None,
                "transformation_type": "Direct Mapping",
                "transformation_logic": row[5].strip() or None,
                "source_role": "value",
            })
    elif kind == "cloud_sheet":
        table = rows[0][1].strip()
        header = next(i for i, r in enumerate(rows)
                      if r and r[0].strip().lower() == "column")
        for row in rows[header + 1:]:
            if not row or not row[0].strip():
                continue
            out.append({
                "table": table, "column": row[0].strip(),
                "datatype": row[1].strip() or None,
                "size": row[2].strip() or None,
                "constraint": None, "description": None,
            })
    return out


# --- running one case ------------------------------------------------------

def run_case(name, archive_dir, layout, note):
    layout = validate_layout(layout)
    catalog = build_catalog(str(archive_dir), layout)
    schemas = {k: json.loads((ROOT / "schemas" / v).read_text())
               for k, v in __import__("kinds").schema_files().items()}
    store = RecordStore(schemas, cache_path=None, verbose=False)

    targets = ["DWH_PARTY"]
    cloud_path = find_cloud_workbook(str(archive_dir), layout)
    cloud_index = {}
    if cloud_path:
        wanted = set(targets)
        for prefix in table_prefixes(layout):
            wanted |= {t[len(prefix):] for t in targets if t.startswith(prefix)}
        sheets = {t: s for t, s in cloud_sheets(cloud_path).items() if t in wanted}
        if sheets:
            cloud_index = assemble.build_cloud_index(cloud_path, sheets, store)

    records = []
    for table in targets:
        records += assemble.resolve_table(table, catalog, store, cloud_index,
                                          cloud_stage(layout), table_prefixes(layout))

    coverage = coverage_metrics(records, layout["stages"])
    chains = chain_diagnostics(records, layout["stages"])

    print(f"\n{'=' * 74}\n{name}\n{note}\n{'=' * 74}")
    print(f"  layout stages        : {layout['stages']}")
    print(f"  cloud workbook found : {bool(cloud_path)}")
    print(f"  records              : {len(coverage) and coverage['records']}")
    print(f"  stages actually built: {assemble.stages_in(records)}")
    print(f"  complete_chains      : {coverage['complete_chains']}")
    for stage in layout["stages"]:
        key = f"stage_{stage}"
        if key in coverage:
            print(f"  {key:21}: {coverage[key]}")
    print(f"  chain diagnostics    : complete {chains['complete_chains']}/{chains['chains']}, "
          f"unmatched tails {chains['unmatched_tails']['count']}")

    print("\n  one field, end to end:")
    for entry in records[0]["lineage"]:
        src = entry["sources"][0]["table"] + "." + entry["sources"][0]["column"] \
              if entry["sources"] else "(origin)"
        print(f"    {entry['stage']:<8} {entry['table']}.{entry['column']:<10}"
              f" {str(entry['datatype']):<9} from {src}")

    out = Path(archive_dir) / "out.xlsx"
    write_workbook(records, str(out))
    print(f"\n  spreadsheet written  : {out.name} ({out.stat().st_size} bytes)")
    return records, coverage


def run_case_quiet(archive_dir, layout):
    """Same pipeline, no printing, for the assertions below."""
    import io, contextlib
    with contextlib.redirect_stdout(io.StringIO()):
        return run_case("", archive_dir, layout, "")


def main():
    store_mod.convert = _fake_convert          # the one stubbed thing
    tmp = Path(tempfile.mkdtemp(prefix="hecate_nocloud_"))
    try:
        a = build_archive(tmp / "A", with_cloud_workbook=True)
        b = build_archive(tmp / "B", with_cloud_workbook=False)
        c = build_archive(tmp / "C", with_cloud_workbook=False)

        run_case("CASE A  baseline", a, LAYOUT_WITH_CLOUD,
                 "layout declares cloud, and the workbook is there")
        run_case("CASE B  cloud declared, workbook missing", b, LAYOUT_WITH_CLOUD,
                 "the realistic accident: config says 4 stages, folder has 3")
        run_case("CASE C  no cloud stage in the layout", c, LAYOUT_NO_CLOUD,
                 "an honest three-stage archive")

        # --- what the three cases must keep proving -----------------------
        (ra, ca), (rb, cb), (rc, cc) = (
            run_case_quiet(a, LAYOUT_WITH_CLOUD),
            run_case_quiet(b, LAYOUT_WITH_CLOUD),
            run_case_quiet(c, LAYOUT_NO_CLOUD))

        checks = [
            ("a missing cloud workbook does not fail the run",
             len(rb) == 2),
            ("a missing cloud workbook does not fail emit",
             True),
            ("the chain simply ends at dwh, with no empty cloud entry",
             [e["stage"] for e in rb[0]["lineage"]] == ["source", "staging", "dwh"]),
            # offline_path is an absolute path, so it necessarily differs
            # between two archives in different folders. Everything that
            # describes the FIELD must match.
            ("records are identical whether or not the layout claims a cloud stage",
             [{k: v for k, v in e.items() if k != "offline_path"}
              for e in rb[0]["lineage"]]
             == [{k: v for k, v in e.items() if k != "offline_path"}
                 for e in rc[0]["lineage"]]),
            ("a three-stage layout reports its chains complete",
             cc["complete_chains"].startswith("2/2")),
            ("BUT a four-stage layout with no cloud workbook reports 0%",
             cb["complete_chains"].startswith("0/2")),
            ("the declared-but-absent stage is reported as 0, not omitted",
             cb.get("stage_cloud", "").startswith("0/2")),
            ("with a cloud workbook present the chain reaches it",
             [e["stage"] for e in ra[0]["lineage"]][-1] == "cloud"),
        ]
        print(f"\n{'=' * 74}\nASSERTIONS\n{'=' * 74}")
        failed = 0
        for label, ok in checks:
            print(f"  {'PASS' if ok else 'FAIL'}  {label}")
            failed += not ok
        print(f"\n  {len(checks) - failed}/{len(checks)} passed")
        return 1 if failed else 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main() or 0)
