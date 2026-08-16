"""Write the assembled dictionary in the hand-built layout.

The JSON is the machine artefact; this is the one a person opens next to the
manual sheet and compares row by row. Same four stage groups, same tail
columns, same order.
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from assemble import stages_in

# Fill colour per known stage; an archive with stages we have no colour for
# still emits, cycling the palette rather than failing.
STAGE_COLOURS = {"source": "FFE8D6", "staging": "D6E8F5",
                 "dwh": "DDEEDD", "cloud": "EADCF0"}
PALETTE = ["FFE8D6", "D6E8F5", "DDEEDD", "EADCF0", "F5E6CC", "E0E0F0"]


def stage_groups(stages) -> list:
    """(display name, stage, fill) per stage, in pipeline order."""
    return [(stage.replace("_", " ").title(), stage,
             STAGE_COLOURS.get(stage, PALETTE[i % len(PALETTE)]))
            for i, stage in enumerate(stages)]
STAGE_COLS = ["Tên Bảng", "Tên Cột", "Đường Dẫn", "datatype", "size"]
TAIL_COLS = ["Mô Tả", "transformation logic", "Logic Notes", "Join / Depends-on"]
WIDTHS = {1: 22, 2: 22, 3: 46, 4: 11, 5: 7, 6: 26, 7: 24, 8: 46, 9: 11, 10: 7,
          11: 20, 12: 18, 13: 48, 14: 11, 15: 7, 16: 20, 17: 18, 18: 60, 19: 11,
          20: 7, 21: 30, 22: 20, 23: 44, 24: 40}
TABLE_TOKEN = re.compile(r"\b(?:STG|DWH|SRC)_[A-Z0-9_]{3,}\b", re.I)


def _write_rows(ws, rows, start_row, should_wrap):
    """Values into cells, formatted as text so Excel cannot reinterpret them.

    Text format matters: without it Excel turns a value like "1-1" into a date,
    which silently corrupts a dictionary.
    """
    for r, row in enumerate(rows, start=start_row):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c, value if value not in ("", None) else None)
            cell.alignment = Alignment(vertical="top", wrap_text=should_wrap(c))
            if isinstance(value, str):
                cell.number_format = "@"


def _apply_layout(ws, widths, freeze):
    for c, width in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = freeze


def _short(path):
    """Archive-relative path, matching how the manual sheet cites evidence."""
    if not path:
        return None
    marker = "/Archive/"
    return "Archive/" + path.split(marker, 1)[1] if marker in path else path


def _tail(record):
    """(description, transformation type, rule, dependencies) for one row."""
    produced = None
    for entry in record["lineage"]:
        if entry["sources"]:
            produced = entry
    source = produced["sources"][0] if produced and produced["sources"] else {}
    logic = source.get("transformation_logic") or ""
    named = {t.upper() for t in TABLE_TOKEN.findall(logic)}
    named -= {(source.get("table") or "").upper()}
    return (record.get("description"), source.get("transformation_type"),
            logic or None, ", ".join(sorted(named)) or None)


def to_rows(lineage_records, groups) -> list:
    rows = []
    for record in lineage_records:
        by_stage = {e["stage"]: e for e in record["lineage"]}
        row = []
        for _, stage, _ in groups:
            entry = by_stage.get(stage)
            row += ([entry["table"], entry["column"], _short(entry["offline_path"]),
                     entry["datatype"], entry["size"]] if entry else [None] * 5)
        rows.append(row + list(_tail(record)))
    return rows


def write_workbook(lineage_records, path, stages=None):
    groups = stage_groups(stages or stages_in(lineage_records))
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, (name, _, color) in enumerate(groups):
        off = i * 5
        ws.cell(2, off + 1, name)
        fill = PatternFill("solid", fgColor=color)
        for k, header in enumerate(STAGE_COLS):
            ws.cell(3, off + k + 1, header)
            ws.cell(2, off + k + 1).fill = fill
            ws.cell(3, off + k + 1).fill = fill
            ws.cell(3, off + k + 1).font = Font(bold=True)
        ws.cell(2, off + 1).font = Font(bold=True)
    tail_start = len(groups) * len(STAGE_COLS) + 1
    for k, header in enumerate(TAIL_COLS):
        cell = ws.cell(3, tail_start + k, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")
    _write_rows(ws, to_rows(lineage_records, groups), 4,
                lambda c: c >= tail_start)
    _apply_layout(ws, WIDTHS, "A4")
    wb.save(path)
    return path


# ------------------------------------------------------- SQL view workbook
#
# Views carry no datatype or size, and only two stages, so they get their own
# column layout rather than a squeezed version of the four-stage one. Matches
# the hand-built SQL ground truth column for column.

VIEW_COLS = ["Source Schema", "Source Table", "Source Column", "View Name",
             "View Column", "Role", "Transformation", "Đường dẫn", "Alias",
             "Ultimate Source", "Via"]
VIEW_WIDTHS = {1: 14, 2: 26, 3: 30, 4: 26, 5: 30, 6: 9, 7: 62, 8: 26, 9: 10,
               10: 30, 11: 34}


def view_rows(view_records) -> list:
    rows = []
    for record in view_records:
        by_stage = {e["stage"]: e for e in record["lineage"]}
        view = by_stage.get("view", {})
        source = (view.get("sources") or [{}])[0]
        resolved = record.get("resolved_source") or {}
        ultimate = ""
        if resolved.get("column"):
            ultimate = ".".join(x for x in (resolved.get("table"),
                                            resolved.get("column")) if x)
        rows.append([
            source.get("schema"), source.get("table"), source.get("column"),
            view.get("table"), view.get("column"), source.get("role"),
            source.get("transformation_logic"),
            Path(view.get("offline_path") or "").name or None,
            source.get("alias"),
            ultimate or None, " -> ".join(resolved.get("via", [])) or None,
        ])
    return rows


def write_view_workbook(view_records, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for j, header in enumerate(VIEW_COLS, start=1):
        cell = ws.cell(1, j, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")
    _write_rows(ws, view_rows(view_records), 2, lambda c: c == 7)
    _apply_layout(ws, VIEW_WIDTHS, "A2")
    wb.save(path)
    return path
